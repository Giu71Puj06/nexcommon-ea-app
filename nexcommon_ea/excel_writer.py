from pathlib import Path
from datetime import datetime, date

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from nexcommon_ea.layout_modulo import mappa_colonne, prima_riga_dati

TEMPLATE = Path(__file__).resolve().parents[1] / "data" / "template_modulo_its.xlsx"


def _to_excel_num(value):
    if value is None or value == "":
        return ""
    try:
        return round(float(value), 3)
    except Exception:
        return value


def _to_excel_date(value):
    """
    Converte la data prova in una vera data Excel.
    La cella verra poi formattata come GG/MM/AAAA.
    Accetta:
    - datetime/date Python
    - stringhe ISO tipo 2024-06-03, 2024-06-03T10:20:30
    - stringhe italiane tipo 03/06/2024
    - stringhe con ora tipo 03/06/2024 10:20:30
    """
    if value is None or value == "":
        return ""

    if isinstance(value, datetime):
        return value.date()

    if isinstance(value, date):
        return value

    text = str(value).strip()
    if not text:
        return ""

    # Rimuove timezone ISO e separatore T, se presenti.
    normalized = text.replace("Z", "").replace("T", " ")

    # Prova formati frequenti.
    formats = [
        "%Y-%m-%d %H:%M:%S.%f",
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d",
        "%d/%m/%Y %H:%M:%S",
        "%d/%m/%Y",
        "%d-%m-%Y %H:%M:%S",
        "%d-%m-%Y",
    ]

    for fmt in formats:
        try:
            return datetime.strptime(normalized[:26], fmt).date()
        except Exception:
            pass

    # Ultimo tentativo per ISO non standard accettato da Python.
    try:
        return datetime.fromisoformat(normalized).date()
    except Exception:
        return text


def _find_or_next_row(ws, matricola: str, col_matricola: int = 4,
                      prima_riga: int = 3) -> int:
    matricola = str(matricola or "").strip()
    candidate_rows = (list(range(prima_riga, prima_riga + 20))
                      + list(range(prima_riga + 31, prima_riga + 51)))

    if matricola:
        for row in candidate_rows:
            value = ws.cell(row=row, column=col_matricola).value
            if str(value).strip() == matricola:
                return row

    for row in candidate_rows:
        if not ws.cell(row=row, column=col_matricola).value:
            return row

    return candidate_rows[-1]


def create_excel_from_summary(summary: dict, output_path: str | Path, template_path: str | Path | None = None) -> Path:
    template = Path(template_path) if template_path else TEMPLATE
    wb = load_workbook(template)
    ws = wb["MODULO CONSEGNA"] if "MODULO CONSEGNA" in wb.sheetnames else wb.active

    # Le colonne vengono lette dalle intestazioni: il modulo ITS ha layout
    # diversi fra i laboratori (vedi layout_modulo.py).
    col, ripiego = mappa_colonne(ws)
    riga_iniziale = prima_riga_dati(ws, col)
    row = _find_or_next_row(ws, summary.get("matricola", ""),
                            col["matricola"], riga_iniziale)

    def scrivi(campo, valore):
        if campo in col and valore not in (None, ""):
            ws.cell(row=row, column=col[campo]).value = valore

    ws.cell(row=row, column=col["lab"]).value = summary.get("laboratorio", 12)

    # Data prova: scritta come vera data Excel e visualizzata GG/MM/AAAA.
    if summary.get("data_prova"):
        data_excel = _to_excel_date(summary.get("data_prova"))
        ws.cell(row=row, column=col["data"]).value = data_excel
        if isinstance(data_excel, date):
            ws.cell(row=row, column=col["data"]).number_format = "DD/MM/YYYY"

    scrivi("matricola", summary.get("matricola"))
    scrivi("localita", summary.get("localita"))

    # Provincia: quella di INSTALLAZIONE (master o record BD), non quella
    # di immatricolazione che compare nel nome del file Vallen.
    scrivi("provincia", summary.get("provincia"))
    scrivi("cliente", summary.get("cliente"))

    # L'ora di inizio prova del modulo e' l'ora di arrivo in sito (par. (*)
    # in calce al modulo), che NON coincide con l'inizio della
    # pressurizzazione e non e' ricavabile dai dati Vallen: la cella resta
    # da compilare e l'orario misurato finisce nelle note.

    scrivi("pressione_inizio", _to_excel_num(summary.get("pressione_inizio_bar")))
    scrivi("pressione_fine", _to_excel_num(summary.get("pressione_fine_bar")))

    gamma = summary.get("gamma_max")
    if gamma is not None:
        scrivi("y_max", _to_excel_num(gamma))

    # Esito: la classificazione resta dell'operatore. Se la classe non e'
    # stata confermata si scrive la proposta, marcata come tale.
    classe = summary.get("classe")
    if classe:
        esito = "IDONEO" if str(classe) == "1" else "NON IDONEO"
    elif summary.get("classe_proposta"):
        etichette = {"0": "NON CLASSIFICABILE", "1": "IDONEO",
                     "2": "NON IDONEO"}
        esito = "DA CONFERMARE: " + etichette.get(
            str(summary["classe_proposta"]), "")
    elif gamma is None:
        esito = "Y MAX DA BD/LISTATO"
    else:
        esito = "DA CONFERMARE"
    if "esito" in ripiego:
        # Il modulo non ha una colonna Esito: il giudizio va nelle note.
        esito_in_nota = esito
    else:
        esito_in_nota = None
        scrivi("esito", esito)

    scrivi("riserva", summary.get("in_sostituzione_di"))
    scrivi("lotto", summary.get("lotto"))

    # A1-A4: differenze della Calibration Table fra verifica di
    # funzionalita' finale e iniziale (Appendice D tab. D6, campi 22-25).
    for campo in ("a1", "a2", "a3", "a4"):
        if summary.get(campo) is not None:
            ws.cell(row=row, column=col[campo]).value = summary[campo]

    note = []

    if esito_in_nota:
        note.append(esito_in_nota)

    if summary.get("ora_inizio_pressurizzazione"):
        note.append("inizio pressurizzazione "
                    + str(summary["ora_inizio_pressurizzazione"])
                    + " (ora di inizio prova da inserire)")

    if summary.get("valutazione_sintesi"):
        note.append(summary["valutazione_sintesi"])

    if summary.get("gradiente_bar_min") is not None:
        note.append(f"grad. {summary['gradiente_bar_min']:.3f} bar/min")

    if summary.get("hits_fondo_finale") is not None:
        note.append(f"FF hits {summary['hits_fondo_finale']}")

    if summary.get("rms_fondo_finale_uv") is not None:
        note.append(f"RMS FF {summary['rms_fondo_finale_uv']:.2f} uV")

    fonte_p = summary.get("fonte_pressione")
    if fonte_p and "IP1" not in str(fonte_p):
        note.append(f"pressioni da {fonte_p}")

    if gamma is not None and summary.get("gamma_source"):
        note.append(f"Ymax da {summary['gamma_source']}")
    elif gamma is None:
        note.append("Ymax assente nel PRIDB: caricare listato/BD o inserirlo a mano")

    stato = summary.get("anagrafica_stato")
    if stato and "non trovata" in str(stato):
        note.append("matricola non in anagrafica: Prov./Localita/Cliente da inserire")

    if summary.get("a1") is not None:
        note.append(f"A1-A4 da {summary.get('a_source', 'Calibration Table')}")
    else:
        note.append("A1-A4: colpi di pulsatore non ricostruibili")

    if summary.get("a_discrepanza"):
        note.append(summary["a_discrepanza"])

    if summary.get("verifica_funzionalita") == "non accettabile":
        note.append("verifica di funzionalita finale NON conforme (par. 19)")

    ripiego_scritte = [c for c in ripiego if c != "esito"]
    if ripiego_scritte:
        note.append("colonne non riconosciute dalle intestazioni: "
                    + ", ".join(ripiego_scritte))

    ws.cell(row=row, column=col["note"]).value = " | ".join(note)

    # Foglio tecnico con tutti i dati estratti, per audit e debugging.
    if "Dati Vallen" in wb.sheetnames:
        del wb["Dati Vallen"]

    audit = wb.create_sheet("Dati Vallen")
    audit.append(["Campo", "Valore"])

    for key, value in summary.items():
        if key == "data_prova":
            parsed_date = _to_excel_date(value)
            audit.append([key, parsed_date])
            if isinstance(parsed_date, date):
                audit.cell(row=audit.max_row, column=2).number_format = "DD/MM/YYYY"
        elif isinstance(value, (list, tuple, dict)):
            audit.append([key, str(value)])
        else:
            audit.append([key, value])

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    thin = Side(style="thin", color="B7B7B7")

    for cell in audit[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.border = Border(bottom=thin)

    audit.column_dimensions["A"].width = 35
    audit.column_dimensions["B"].width = 45

    output_path = Path(output_path)
    wb.save(output_path)
    return output_path
