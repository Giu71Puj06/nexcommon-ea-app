"""
Anagrafica serbatoi (file master Excel).

Il pacchetto Vallen contiene i dati della PROVA (pressioni, hit, marker),
ma non i dati ANAGRAFICI del serbatoio (provincia e localita di
installazione, gasista/cliente) ne il valore Y Max/Gamma registrato
dall'operatore. Questi vivono nel registro master (es.
'Master_L_autogas_..._elab.xlsx').

Questo modulo legge il master, costruisce un indice per numero di
matricola e permette di recuperare i campi mancanti da inserire nel
modulo ITS.

Il match avviene per matricola (colonna 'N°matr'). Le colonne sono
individuate per NOME di intestazione, non per posizione, cosi funziona
anche se l'ordine cambia tra i fogli.
"""

import re
from pathlib import Path

from openpyxl import load_workbook

# Mappa: nome campo interno -> possibili intestazioni nel master.
COLUMN_ALIASES = {
    "matricola": ["n°matr", "n matr", "nmatr", "matricola"],
    "provincia": ["prov. inst", "prov inst", "provincia inst", "prov. installazione"],
    "localita": ["loc. inst", "loc inst", "localita inst", "localita installazione"],
    "cliente": ["proprietario", "gasista", "cliente"],
    "gamma_max": ["y", "y max", "ymax", "gamma", "gamma max"],
    "pr_fab": ["pr. fab", "pr fab", "prov. fab", "provincia fabbricazione"],
    "lotto": ["sigla lotto", "lotto"],
    "matr_sost": ["matr. sost.", "matr sost", "matricola sostituita"],
    "anno_matr": ["anno matr", "anno matricola"],
    "esito": ["esito prova", "esito"],
    "fabbricante": ["fabbricante"],
    "rivestimento": ["rivest.", "rivestimento"],
}


def normalize_matricola(value) -> str:
    """
    Normalizza una matricola per il confronto: toglie spazi, il '.0'
    dei float Excel e gli zeri non significativi quando e tutta numerica.
    """
    if value is None:
        return ""
    s = str(value).strip()
    if not s:
        return ""
    s = re.sub(r"\.0$", "", s)
    if s.isdigit():
        return str(int(s))  # rimuove eventuali zeri iniziali
    return s.upper()


def _header_map(header_row) -> dict:
    """Da riga di intestazione a {campo_interno: indice_colonna}."""
    idx = {}
    for i, cell in enumerate(header_row):
        if cell is None:
            continue
        name = str(cell).strip().lower()
        for field, aliases in COLUMN_ALIASES.items():
            if field in idx:
                continue
            if name in aliases:
                idx[field] = i
                break
    return idx


def load_registry(path: str | Path) -> dict:
    """
    Carica il master e ritorna {matricola_normalizzata: record}.

    Un record contiene i campi definiti in COLUMN_ALIASES effettivamente
    presenti. Se una matricola compare in piu fogli, si preferisce il
    record con il valore Y (gamma) compilato, poi il primo trovato.
    """
    path = Path(path)
    wb = load_workbook(path, read_only=True, data_only=True)
    registry: dict[str, dict] = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = ws.iter_rows(values_only=True)
        try:
            header = next(rows)
        except StopIteration:
            continue
        cols = _header_map(header)
        if "matricola" not in cols:
            continue  # non e un foglio anagrafico

        for row in rows:
            matr_raw = row[cols["matricola"]] if cols["matricola"] < len(row) else None
            key = normalize_matricola(matr_raw)
            if not key:
                continue

            record = {"_foglio": sheet_name, "matricola": key}
            for field, ci in cols.items():
                if field == "matricola":
                    continue
                value = row[ci] if ci < len(row) else None
                if value is not None and str(value).strip() != "":
                    record[field] = value

            existing = registry.get(key)
            if existing is None:
                registry[key] = record
            else:
                # Preferisci il record che ha il gamma (Y) compilato.
                if "gamma_max" in record and "gamma_max" not in existing:
                    registry[key] = record

    wb.close()
    return registry


def lookup(registry: dict, matricola) -> dict | None:
    """Cerca una matricola nell'anagrafica. Ritorna il record o None."""
    if not registry:
        return None
    return registry.get(normalize_matricola(matricola))


def enrich_summary(summary: dict, record: dict | None) -> dict:
    """
    Integra il summary della prova con i dati anagrafici del master.

    Non sovrascrive un valore gia presente e valido (es. gamma da
    listato Vallen ha priorita su Y del master). Segnala l'esito del
    match in 'anagrafica_stato'.
    """
    if not record:
        summary["anagrafica_stato"] = "matricola non trovata in anagrafica"
        return summary

    # Provincia e localita: il master e la fonte corretta (installazione),
    # quindi hanno priorita sull'inferenza dal nome file.
    if record.get("provincia"):
        summary["provincia"] = str(record["provincia"]).strip()
    if record.get("localita"):
        summary["localita"] = str(record["localita"]).strip()
    if record.get("cliente"):
        summary["cliente"] = str(record["cliente"]).strip()

    # Y Max/Gamma dal master solo se non gia ottenuto da listato/BD/manuale.
    if summary.get("gamma_max") is None and record.get("gamma_max") is not None:
        try:
            summary["gamma_max"] = float(str(record["gamma_max"]).replace(",", "."))
            summary["gamma_source"] = "anagrafica master (colonna Y)"
        except (TypeError, ValueError):
            pass

    # Dati di supporto utili nel modulo.
    if record.get("matr_sost"):
        summary["in_sostituzione_di"] = str(record["matr_sost"]).strip()
    if record.get("lotto"):
        summary["lotto_master"] = str(record["lotto"]).strip()
    if record.get("rivestimento"):
        summary["rivestimento"] = str(record["rivestimento"]).strip()

    summary["anagrafica_stato"] = f"trovata in '{record.get('_foglio', '?')}'"
    return summary
